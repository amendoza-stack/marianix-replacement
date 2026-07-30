import os
import csv
from datetime import date
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class ReportGenerators:
    
    # -------------------------------------------------------------------------
    # 1. Reporte 331 (Texto Plano Normativo de Presentación)
    # -------------------------------------------------------------------------
    @staticmethod
    def generate_reporte_331_txt(recetas_data: List[Dict[str, Any]], output_path: str):
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("HDR331|PRESENTACION_LIQUIDACION_OFICIAL|2026\n")
            for r in recetas_data:
                cuit_farm = str(r.get("farmacia_cuit", "")).zfill(11)
                num_rec = str(r.get("numero_receta", "")).ljust(15)
                f_disp = str(r.get("fecha_dispensa", "")).replace("-", "")[0:8]
                dni_afi = str(r.get("afiliado_dni", "")).zfill(8)
                pvp = f"{r.get('total_pvp', 0.0):.2f}".zfill(10)
                cob = f"{r.get('total_cobertura_os', 0.0):.2f}".zfill(10)
                
                line = f"331|{cuit_farm}|{num_rec}|{f_disp}|{dni_afi}|{pvp}|{cob}\n"
                f.write(line)
            f.write(f"TRL331|TOTAL_REGISTROS|{len(recetas_data)}\n")

    # -------------------------------------------------------------------------
    # 2. Reporte Excel con Estilos (openpyxl)
    # -------------------------------------------------------------------------
    @staticmethod
    def generate_excel_report(recetas_data: List[Dict[str, Any]], output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte por Farmacia"

        # Estilos Excel
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_font = Font(name="Arial", size=14, bold=True, color="1F4E78")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Título
        ws.merge_cells("A1:H1")
        ws["A1"] = "REPORTE DETALLADO DE LIQUIDACION POR FARMACIA"
        ws["A1"].font = title_font
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 30

        # Encabezados de Tabla
        headers = ["N° Receta", "Fecha Dispensa", "Farmacia", "Afiliado DNI", "Médico", "PVP Total ($)", "Cobertura OS ($)", "Monto Bonificado ($)"]
        ws.append([]) # Línea vacía
        ws.append(headers)

        ws.row_dimensions[3].height = 25
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        # Datos
        row_idx = 4
        for r in recetas_data:
            ws.append([
                r.get("numero_receta"),
                str(r.get("fecha_dispensa")),
                r.get("farmacia_nombre"),
                r.get("afiliado_dni"),
                r.get("medico_nombre"),
                r.get("total_pvp"),
                r.get("total_cobertura_os"),
                r.get("monto_bonificado")
            ])
            for col_idx in range(1, 9):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = thin_border
                if col_idx in [6, 7, 8]:
                    c.number_format = "$#,##0.00"
                    c.alignment = align_right
            row_idx += 1

        # Ajuste automático de ancho de columnas
        for col_idx in range(1, 9):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in range(3, ws.max_row + 1):
                val = str(ws.cell(row=row, column=col_idx).value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        wb.save(output_path)

    # -------------------------------------------------------------------------
    # 3. Reporte PDF con Tablas (ReportLab)
    # -------------------------------------------------------------------------
    @staticmethod
    def generate_pdf_report(recetas_data: List[Dict[str, Any]], output_path: str):
        doc = SimpleDocTemplate(output_path, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=16, leading=20, textColor=colors.HexColor('#1F4E78'))
        elements.append(Paragraph("Reporte Consolidado de Liquidaciones de Farmacia", title_style))
        elements.append(Spacer(1, 15))

        table_data = [["N° Receta", "Fecha Disp.", "Farmacia", "Afiliado DNI", "Médico", "PVP Total", "Cobertura OS", "Bonificado"]]

        for r in recetas_data:
            table_data.append([
                r.get("numero_receta", ""),
                str(r.get("fecha_dispensa", "")),
                r.get("farmacia_nombre", "")[:20],
                r.get("afiliado_dni", ""),
                r.get("medico_nombre", "")[:20],
                f"",
                f"",
                f""
            ])

        t = Table(table_data, colWidths=[80, 75, 130, 80, 130, 80, 80, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ]))

        elements.append(t)
        doc.build(elements)

    # -------------------------------------------------------------------------
    # 4. Reporte CSV Estándar
    # -------------------------------------------------------------------------
    @staticmethod
    def generate_csv_report(recetas_data: List[Dict[str, Any]], output_path: str):
        headers = ["numero_receta", "fecha_dispensa", "farmacia_nombre", "afiliado_dni", "medico_nombre", "total_pvp", "total_cobertura_os", "monto_bonificado"]
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for r in recetas_data:
                writer.writerow({
                    "numero_receta": r.get("numero_receta"),
                    "fecha_dispensa": r.get("fecha_dispensa"),
                    "farmacia_nombre": r.get("farmacia_nombre"),
                    "afiliado_dni": r.get("afiliado_dni"),
                    "medico_nombre": r.get("medico_nombre"),
                    "total_pvp": r.get("total_pvp"),
                    "total_cobertura_os": r.get("total_cobertura_os"),
                    "monto_bonificado": r.get("monto_bonificado")
                })